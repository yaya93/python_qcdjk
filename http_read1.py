#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @FileName  :http_read1.py
# @Time      :2020/7/3 22:36
# @Author    :yaya
from openpyxl import load_workbook
def read_data(file_name,sheet_name):
    wb=load_workbook(file_name)
    sheet=wb[sheet_name]
    all_case=[]
    for i in range(2,sheet.max_row):
        case=[]
        for j in range(1,sheet.max_column-1):
            case.append( sheet.cell(row=i, column=j).value)
        all_case.append(case)
    return all_case
if __name__ == '__main__':

    all_case=read_data('login.xlsx','recharge')
    print('所有的数据为:',all_case)